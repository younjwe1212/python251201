import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
                             QLabel, QProgressBar, QHeaderView, QMessageBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class CrawlerThread(QThread):
    """크롤링을 백그라운드에서 실행하는 스레드"""
    progress = pyqtSignal(str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def run(self):
        try:
            self.progress.emit("브라우저를 시작하는 중...")
            
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')  # GUI에서는 headless 모드 사용
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--log-level=3')  # 로그 최소화
            
            driver = webdriver.Chrome(options=options)
            
            self.progress.emit("김프가 웹사이트에 접속 중...")
            driver.get('https://kimpga.com/')
            
            # Wait for the table to load
            wait = WebDriverWait(driver, 10)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            
            self.progress.emit("데이터를 추출하는 중...")
            
            tables = driver.find_elements(By.TAG_NAME, "table")
            
            target_table = None
            for table in tables:
                if "이름" in table.text:
                    target_table = table
                    break
            
            if not target_table and tables:
                target_table = tables[0]
                
            coins = []
            
            if target_table:
                rows = target_table.find_elements(By.TAG_NAME, "tr")
                
                for row in rows:
                    if row.find_elements(By.TAG_NAME, "th"):
                        continue
                        
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        continue
                    
                    try:
                        # Cell 0: Name/Symbol
                        c1 = cells[0]
                        lines = c1.text.split('\n')
                        
                        name = ""
                        symbol = ""
                        
                        if len(lines) >= 2:
                            name = lines[0]
                            symbol = lines[1]
                        else:
                            name = c1.text
                        
                        # Cell 1: Price
                        price = "N/A"
                        if len(cells) > 1:
                            price_text = cells[1].text
                            price = price_text.split('\n')[0]
                        
                        name = name.strip()
                        symbol = symbol.strip()
                        price = price.strip()
                        
                        if name and price:
                            coins.append({
                                'name': name,
                                'symbol': symbol,
                                'price': price
                            })
                            
                        if len(coins) >= 20:
                            break
                            
                    except Exception:
                        continue
            
            driver.quit()
            
            self.progress.emit(f"완료! {len(coins)}개의 코인 데이터를 가져왔습니다.")
            self.finished.emit(coins)
            
        except Exception as e:
            self.error.emit(f"오류 발생: {str(e)}")


class KimpgaGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.coins = []  # 크롤링된 코인 데이터 저장
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('김프가 코인 크롤러')
        self.setGeometry(100, 100, 900, 600)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        layout = QVBoxLayout()
        central_widget.setLayout(layout)
        
        # 제목
        title = QLabel('김프가 상위 20개 코인 정보')
        title.setFont(QFont('맑은 고딕', 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # 버튼 레이아웃
        button_layout = QHBoxLayout()
        
        self.crawl_button = QPushButton('데이터 가져오기')
        self.crawl_button.setFont(QFont('맑은 고딕', 11))
        self.crawl_button.setMinimumHeight(40)
        self.crawl_button.clicked.connect(self.start_crawling)
        button_layout.addWidget(self.crawl_button)
        
        self.clear_button = QPushButton('초기화')
        self.clear_button.setFont(QFont('맑은 고딕', 11))
        self.clear_button.setMinimumHeight(40)
        self.clear_button.clicked.connect(self.clear_table)
        button_layout.addWidget(self.clear_button)
        
        self.export_button = QPushButton('Excel로 저장')
        self.export_button.setFont(QFont('맑은 고딕', 11))
        self.export_button.setMinimumHeight(40)
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setEnabled(False)  # 초기에는 비활성화
        button_layout.addWidget(self.export_button)
        
        layout.addLayout(button_layout)
        
        # 상태 레이블
        self.status_label = QLabel('데이터 가져오기 버튼을 클릭하세요.')
        self.status_label.setFont(QFont('맑은 고딕', 10))
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        # 프로그레스 바
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # 테이블
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['순위', '이름', '심볼', '현재가'])
        self.table.setFont(QFont('맑은 고딕', 10))
        
        # 테이블 헤더 스타일
        header = self.table.horizontalHeader()
        header.setFont(QFont('맑은 고딕', 10, QFont.Bold))
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        
        # 테이블 설정
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        
        layout.addWidget(self.table)
        
        # 스타일시트 적용
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            QPushButton#exportButton {
                background-color: #2196F3;
            }
            QPushButton#exportButton:hover {
                background-color: #0b7dda;
            }
            QPushButton#exportButton:pressed {
                background-color: #0969c3;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
            }
            QLabel {
                color: #333;
            }
        """)
        
        # Export 버튼에 ID 설정
        self.export_button.setObjectName("exportButton")
        
    def start_crawling(self):
        """크롤링 시작"""
        self.crawl_button.setEnabled(False)
        self.clear_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # 무한 프로그레스
        self.status_label.setText("크롤링 시작...")
        
        # 크롤러 스레드 시작
        self.crawler_thread = CrawlerThread()
        self.crawler_thread.progress.connect(self.update_status)
        self.crawler_thread.finished.connect(self.display_results)
        self.crawler_thread.error.connect(self.display_error)
        self.crawler_thread.start()
        
    def update_status(self, message):
        """상태 업데이트"""
        self.status_label.setText(message)
        
    def display_results(self, coins):
        """결과를 테이블에 표시"""
        self.coins = coins  # 데이터 저장
        self.table.setRowCount(len(coins))
        
        for i, coin in enumerate(coins):
            # 순위
            rank_item = QTableWidgetItem(str(i + 1))
            rank_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 0, rank_item)
            
            # 이름
            name_item = QTableWidgetItem(coin['name'])
            self.table.setItem(i, 1, name_item)
            
            # 심볼
            symbol_item = QTableWidgetItem(coin['symbol'])
            symbol_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 2, symbol_item)
            
            # 가격
            price_item = QTableWidgetItem(coin['price'])
            price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(i, 3, price_item)
        
        self.progress_bar.setVisible(False)
        self.crawl_button.setEnabled(True)
        self.clear_button.setEnabled(True)
        self.export_button.setEnabled(True)  # Excel 저장 버튼 활성화
        
    def display_error(self, error_message):
        """에러 표시"""
        self.status_label.setText(error_message)
        self.progress_bar.setVisible(False)
        self.crawl_button.setEnabled(True)
        self.clear_button.setEnabled(True)
        
    def export_to_excel(self):
        """Excel 파일로 저장"""
        if not self.coins:
            QMessageBox.warning(self, '경고', '저장할 데이터가 없습니다.')
            return
        
        try:
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "코인 정보"
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 작성
            headers = ['순위', '이름', '심볼', '현재가']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 데이터 작성
            for i, coin in enumerate(self.coins, 2):
                ws.cell(row=i, column=1, value=i-1).alignment = Alignment(horizontal="center")
                ws.cell(row=i, column=2, value=coin['name'])
                ws.cell(row=i, column=3, value=coin['symbol']).alignment = Alignment(horizontal="center")
                ws.cell(row=i, column=4, value=coin['price']).alignment = Alignment(horizontal="right")
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            
            # 메타데이터 추가 (별도 시트)
            ws_meta = wb.create_sheet("정보")
            ws_meta['A1'] = "크롤링 일시"
            ws_meta['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_meta['A2'] = "데이터 출처"
            ws_meta['B2'] = "https://kimpga.com/"
            ws_meta['A3'] = "코인 개수"
            ws_meta['B3'] = len(self.coins)
            
            # 파일 저장
            filename = "coin_results.xlsx"
            wb.save(filename)
            
            QMessageBox.information(self, '성공', f'{filename} 파일이 저장되었습니다.\n저장 위치: {filename}')
            self.status_label.setText(f'Excel 파일 저장 완료: {filename}')
            
        except Exception as e:
            QMessageBox.critical(self, '오류', f'Excel 파일 저장 중 오류가 발생했습니다:\n{str(e)}')
        
    def clear_table(self):
        """테이블 초기화"""
        self.table.setRowCount(0)
        self.coins = []  # 데이터 초기화
        self.status_label.setText('데이터 가져오기 버튼을 클릭하세요.')
        self.export_button.setEnabled(False)  # Excel 저장 버튼 비활성화


if __name__ == '__main__':
    app = QApplication(sys.argv)
    gui = KimpgaGUI()
    gui.show()
    sys.exit(app.exec_())
